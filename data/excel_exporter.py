"""
Exports everything to a formatted Excel workbook.

Sheets:
  1. Facebook Groups — every group found with status, city, URL
  2. Posts Made — history of what was posted where
  3. Reddit Subreddits — Auburn-relevant subs to post in
  4. Reddit Posts — specific posts to comment on/DM authors
  5. DM Conversations — funnel tracking
  6. Stats Dashboard — summary numbers
"""

import asyncio
from datetime import datetime
from pathlib import Path
import aiosqlite
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path(__file__).parent.parent / "exports"
OUTPUT_DIR.mkdir(exist_ok=True)

# Auburn colors
AUBURN_ORANGE = "E87722"
AUBURN_BLUE = "03244D"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
GREEN = "C8E6C9"
RED = "FFCDD2"


def header_style(cell, bg=AUBURN_ORANGE, fg=WHITE):
    cell.font = Font(bold=True, color=fg, size=11)
    cell.fill = PatternFill(fill_type="solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_row(ws, row_num: int, num_cols: int, fill_color: str = None):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if fill_color:
            cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
        border = Border(
            bottom=Side(style="thin", color="DDDDDD"),
        )
        cell.border = border


def auto_column_widths(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


async def build_facebook_sheet(wb: openpyxl.Workbook):
    ws = wb.active
    ws.title = "Facebook Groups"
    ws.row_dimensions[1].height = 30

    headers = ["#", "Group Name", "URL", "City", "State", "Members", "Posted?", "Found At"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        header_style(cell)

    async with aiosqlite.connect(Path(__file__).parent / "auburn_blueprint.db") as db:
        cursor = await db.execute(
            "SELECT id, group_name, group_url, city, state, member_count, posted, found_at FROM groups_found ORDER BY city, posted"
        )
        rows = await cursor.fetchall()

    for i, row in enumerate(rows, 2):
        group_id, name, url, city, state, members, posted, found_at = row
        fill = GREEN if posted else LIGHT_GRAY

        ws.cell(row=i, column=1, value=group_id)
        ws.cell(row=i, column=2, value=name)

        # Clickable hyperlink
        url_cell = ws.cell(row=i, column=3, value=url)
        url_cell.hyperlink = url
        url_cell.font = Font(color="0563C1", underline="single")

        ws.cell(row=i, column=4, value=city)
        ws.cell(row=i, column=5, value=state)
        ws.cell(row=i, column=6, value=members)
        ws.cell(row=i, column=7, value="YES" if posted else "NO")
        ws.cell(row=i, column=8, value=found_at)

        style_row(ws, i, len(headers), fill)

    ws.freeze_panes = "A2"
    auto_column_widths(ws)

    # Add filter
    ws.auto_filter.ref = f"A1:H{len(rows)+1}"

    return len(rows)


async def build_posts_sheet(wb: openpyxl.Workbook):
    ws = wb.create_sheet("Posts Made")
    ws.row_dimensions[1].height = 30

    headers = ["Group Name", "Group URL", "Message Preview", "Posted At", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        header_style(cell, bg=AUBURN_BLUE)

    async with aiosqlite.connect(Path(__file__).parent / "auburn_blueprint.db") as db:
        cursor = await db.execute(
            "SELECT group_name, group_url, message_used, posted_at, status FROM posts_made ORDER BY posted_at DESC"
        )
        rows = await cursor.fetchall()

    for i, row in enumerate(rows, 2):
        name, url, msg, posted_at, status = row
        ws.cell(row=i, column=1, value=name)

        url_cell = ws.cell(row=i, column=2, value=url)
        url_cell.hyperlink = url
        url_cell.font = Font(color="0563C1", underline="single")

        ws.cell(row=i, column=3, value=(msg or "")[:150] + "...")
        ws.cell(row=i, column=4, value=posted_at)
        ws.cell(row=i, column=5, value=status)

        fill = GREEN if status == "sent" else RED
        style_row(ws, i, len(headers), fill)

    ws.freeze_panes = "A2"
    auto_column_widths(ws)
    return len(rows)


async def build_dm_sheet(wb: openpyxl.Workbook):
    ws = wb.create_sheet("DM Funnel")
    ws.row_dimensions[1].height = 30

    headers = ["Name", "FB User ID", "Their Last Message", "Our Last Reply", "Funnel Stage", "Converted?", "Last Activity"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        header_style(cell, bg="5C4033")

    async with aiosqlite.connect(Path(__file__).parent / "auburn_blueprint.db") as db:
        cursor = await db.execute(
            "SELECT fb_user_name, fb_user_id, last_message, our_last_reply, stage, converted, last_activity FROM dm_conversations ORDER BY last_activity DESC"
        )
        rows = await cursor.fetchall()

    for i, row in enumerate(rows, 2):
        name, uid, their_msg, our_reply, stage, converted, last_act = row
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=uid)
        ws.cell(row=i, column=3, value=(their_msg or "")[:100])
        ws.cell(row=i, column=4, value=(our_reply or "")[:100])
        ws.cell(row=i, column=5, value=f"Stage {stage}/3")
        ws.cell(row=i, column=6, value="YES ($50)" if converted else "Not yet")
        ws.cell(row=i, column=7, value=last_act)

        fill = GREEN if converted else LIGHT_GRAY
        style_row(ws, i, len(headers), fill)

    ws.freeze_panes = "A2"
    auto_column_widths(ws)
    return len(rows)


def build_reddit_sheet(wb: openpyxl.Workbook, reddit_data: dict):
    """Build Reddit sheets from scan results."""
    # Subreddits sheet
    ws_subs = wb.create_sheet("Reddit Subreddits")
    sub_headers = ["Subreddit", "URL", "Title", "Subscribers", "Description"]
    for col, h in enumerate(sub_headers, 1):
        cell = ws_subs.cell(row=1, column=col, value=h)
        header_style(cell, bg="FF4500")  # Reddit orange

    for i, sub in enumerate(reddit_data.get("subreddits", []), 2):
        ws_subs.cell(row=i, column=1, value=sub.get("name", ""))
        url_cell = ws_subs.cell(row=i, column=2, value=sub.get("url", ""))
        url_cell.hyperlink = sub.get("url", "")
        url_cell.font = Font(color="0563C1", underline="single")
        ws_subs.cell(row=i, column=3, value=sub.get("title", ""))
        ws_subs.cell(row=i, column=4, value=sub.get("subscribers", 0))
        ws_subs.cell(row=i, column=5, value=sub.get("description", ""))
        style_row(ws_subs, i, len(sub_headers), LIGHT_GRAY)

    auto_column_widths(ws_subs)

    # Reddit posts sheet
    ws_posts = wb.create_sheet("Reddit Posts (Engage)")
    post_headers = ["Subreddit", "Title", "URL", "Author", "Score", "Comments"]
    for col, h in enumerate(post_headers, 1):
        cell = ws_posts.cell(row=1, column=col, value=h)
        header_style(cell, bg="FF6314")

    for i, post in enumerate(reddit_data.get("posts", [])[:500], 2):
        ws_posts.cell(row=i, column=1, value=f"r/{post.get('subreddit', '')}")
        ws_posts.cell(row=i, column=2, value=post.get("title", "")[:100])
        url_cell = ws_posts.cell(row=i, column=3, value=post.get("url", ""))
        url_cell.hyperlink = post.get("url", "")
        url_cell.font = Font(color="0563C1", underline="single")
        ws_posts.cell(row=i, column=4, value=post.get("author", ""))
        ws_posts.cell(row=i, column=5, value=post.get("score", 0))
        ws_posts.cell(row=i, column=6, value=post.get("num_comments", 0))
        style_row(ws_posts, i, len(post_headers), LIGHT_GRAY)

    auto_column_widths(ws_posts)


async def build_stats_sheet(wb: openpyxl.Workbook, reddit_data: dict = None):
    ws = wb.create_sheet("Stats Dashboard", 0)

    from data.db import get_stats
    stats = await get_stats()

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20

    title = ws.cell(row=1, column=1, value="Auburn Blueprint — Outreach Dashboard")
    title.font = Font(bold=True, size=16, color=AUBURN_ORANGE)
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 35

    date_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    date_cell.font = Font(italic=True, color="888888")
    ws.merge_cells("A2:B2")

    data_rows = [
        ("", ""),
        ("FACEBOOK", ""),
        ("Groups Found", stats["groups_found"]),
        ("Groups Posted To", stats["groups_posted"]),
        ("Groups Remaining", stats["groups_found"] - stats["groups_posted"]),
        ("", ""),
        ("REDDIT", ""),
        ("Subreddits Discovered", reddit_data.get("total_subreddits", 0) if reddit_data else 0),
        ("Posts to Engage With", reddit_data.get("total_posts", 0) if reddit_data else 0),
        ("", ""),
        ("FUNNEL", ""),
        ("DM Conversations", stats["dm_conversations"]),
        ("Conversions (Paid)", stats["converted"]),
        ("Revenue", f"${stats['converted'] * 50}"),
        ("Conversion Rate", f"{(stats['converted']/stats['dm_conversations']*100):.1f}%" if stats["dm_conversations"] else "N/A"),
        ("", ""),
        ("MARKETPLACE", ""),
        ("Active Listings", stats["active_listings"]),
    ]

    for i, (label, value) in enumerate(data_rows, 4):
        label_cell = ws.cell(row=i, column=1, value=label)
        value_cell = ws.cell(row=i, column=2, value=value)

        if label in ("FACEBOOK", "REDDIT", "FUNNEL", "MARKETPLACE"):
            label_cell.font = Font(bold=True, color=WHITE)
            label_cell.fill = PatternFill(fill_type="solid", fgColor=AUBURN_BLUE)
            value_cell.fill = PatternFill(fill_type="solid", fgColor=AUBURN_BLUE)
        elif label:
            label_cell.font = Font(bold=False)
            value_cell.font = Font(bold=True)
            if i % 2 == 0:
                label_cell.fill = PatternFill(fill_type="solid", fgColor=LIGHT_GRAY)
                value_cell.fill = PatternFill(fill_type="solid", fgColor=LIGHT_GRAY)


async def export_to_excel(reddit_data: dict = None) -> str:
    """Build and save the full Excel workbook. Returns the file path."""
    wb = openpyxl.Workbook()

    # Build stats first (goes to sheet index 0)
    await build_stats_sheet(wb, reddit_data)
    fb_count = await build_facebook_sheet(wb)
    posts_count = await build_posts_sheet(wb)
    dm_count = await build_dm_sheet(wb)

    if reddit_data:
        build_reddit_sheet(wb, reddit_data)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = OUTPUT_DIR / f"auburn_blueprint_outreach_{timestamp}.xlsx"
    wb.save(str(filename))

    return str(filename)
